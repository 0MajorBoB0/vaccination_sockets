from threading import Lock
from flask import Flask, render_template, session, request, \
    copy_current_request_context, redirect, url_for
from flask_socketio import SocketIO, emit, join_room, leave_room, \
    close_room, rooms, disconnect

# STEP 3: DB imports
import os
from sqlalchemy import create_engine, text
from sqlalchemy.pool import QueuePool
from contextlib import contextmanager

# STEP 4: Admin imports
from functools import wraps

# STEP 5: Utility imports
import uuid
import random
import string
import datetime
from datetime import timezone
import time

# Set this variable to "threading", "eventlet" or "gevent" to test the
# different async modes, or leave it set to None for the application to choose
# the best option based on installed packages.
async_mode = None

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret!'
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours in seconds
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = True  # Required for HTTPS (PythonAnywhere)
socketio = SocketIO(app, async_mode=async_mode, logger=True, engineio_logger=True, cors_allowed_origins="*")

# In-memory tracking for duplicate login prevention
# Format: {participant_id: {'browser_token': 'xxx', 'last_activity': timestamp}}
active_participants = {}
active_participants_lock = Lock()
thread = None
thread_lock = Lock()

# STEP 4: Admin configuration
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

# STEP 4: Admin helper functions
def require_admin():
    """Check if current user is logged in as admin."""
    return session.get("admin_ok") is True

def admin_required(f):
    """Decorator to require admin login."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not require_admin():
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated_function


# STEP 5: Database configuration with connection pooling
DB_HOST = os.environ.get("DB_HOST", "GameTheoryUDE26.mysql.eu.pythonanywhere-services.com")
DB_USER = os.environ.get("DB_USER", "GameTheoryUDE26")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "UDE2020EM")
DB_NAME = os.environ.get("DB_NAME", "GameTheoryUDE26$vaccination_game")
DB_PORT = int(os.environ.get("DB_PORT", "3306"))

# Create engine with connection pooling (prevents "too many connections" error)
db_engine = create_engine(
    f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}",
    poolclass=QueuePool,
    pool_size=10,           # Max 10 persistent connections
    max_overflow=20,        # Up to 20 additional connections in peaks
    pool_pre_ping=True,     # Check connection health before using
    pool_recycle=3600,      # Recycle connections after 1 hour
    echo=False,             # Don't log SQL queries
)

@contextmanager
def get_db():
    """Context manager for database connections."""
    conn = db_engine.connect()
    try:
        yield conn
    finally:
        conn.close()


# STEP 5: Utility functions
def create_code(n=6):
    """Generate unique participant code (no confusing chars)."""
    chars = (string.ascii_uppercase + string.digits).replace("O","").replace("0","").replace("I","").replace("1","")
    return "".join(random.choice(chars) for _ in range(n))

def utc_now():
    """Get current UTC time (timezone-aware)."""
    return datetime.datetime.now(timezone.utc).replace(microsecond=0)

def iso_utc(dt):
    """Convert datetime to ISO string with Z suffix."""
    return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def get_participant_state():
    """Get current participant and session state from database."""
    participant_id = session.get("participant_id")
    session_id = session.get("session_id")
    browser_token = session.get("browser_token")

    if not participant_id or not session_id:
        return None, None

    # Update activity timestamp for duplicate-login detection
    if participant_id and browser_token:
        with active_participants_lock:
            if participant_id in active_participants:
                # Only update if browser_token matches (prevent hijacking)
                if active_participants[participant_id]['browser_token'] == browser_token:
                    active_participants[participant_id]['last_activity'] = time.time()

    with get_db() as conn:
        p_result = conn.execute(text("""
            SELECT p.id, p.session_id, p.code, p.joined, p.join_number,
                   p.current_round, p.ptype, p.balance, p.ready_for_next,
                   s.status, s.rounds, s.group_size
            FROM participants p
            JOIN sessions s ON p.session_id = s.id
            WHERE p.id = :pid AND s.id = :sid
        """), {"pid": participant_id, "sid": session_id})

        result = p_result.fetchone()
        if not result:
            return None, None

        participant = dict(result._mapping)
        return participant, session_id

    return None, None


# STEP 8: Game mechanics - TYPE_COST configuration
TYPE_COST = {
    1: {"B": [4, 3, 2, 1, 0],  "A": 4},
    2: {"B": [8, 6, 4, 2, 0],  "A": 4},
    3: {"B": [4, 3, 2, 1, 0],  "A": 8},
    4: {"B": [8, 6, 4, 2, 0],  "A": 8},
    5: {"B": [24, 18, 12, 6, 0], "A": 32},
    6: {"B": [64, 48, 32, 16, 0], "A": 32},
}
B_COLS = 5

def a_cost_for(ptype: int) -> float:
    """Get cost for Option A for a given player type."""
    return TYPE_COST.get(ptype, TYPE_COST[1])["A"]

def b_cost_adapt(ptype: int, others_A: int, N: int) -> float:
    """Calculate cost for Option B based on how many others chose A."""
    if ptype not in TYPE_COST:
        ptype = 1
    b = TYPE_COST[ptype]["B"]
    N = max(1, int(N))
    others_A = max(0, min(int(others_A), max(0, N-1)))
    if N <= 1:
        return float(b[0])
    frac = others_A / float(N - 1)
    x = frac * B_COLS
    col = int(x + 0.5)
    col = max(1, min(B_COLS, col))
    return float(b[col - 1])


# STEP 5: Initialize database schema
def init_db():
    """Initialize database schema with all required tables."""
    print("Initializing database schema...")

    with get_db() as conn:
        # Sessions table
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS sessions (
                id VARCHAR(36) PRIMARY KEY,
                name VARCHAR(255),
                group_size INT,
                rounds INT,
                starting_balance DECIMAL(10,2) DEFAULT 500,
                created_at VARCHAR(30),
                archived TINYINT DEFAULT 0,
                status VARCHAR(20) DEFAULT 'lobby'
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """))

        # Participants table
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS participants (
                id VARCHAR(36) PRIMARY KEY,
                session_id VARCHAR(36),
                code VARCHAR(10) UNIQUE,
                joined TINYINT DEFAULT 0,
                join_number INT,
                current_round INT DEFAULT 1,
                balance DECIMAL(10,2) DEFAULT 0,
                completed TINYINT DEFAULT 0,
                created_at VARCHAR(30),
                ready_for_next TINYINT DEFAULT 0,
                INDEX idx_session (session_id),
                INDEX idx_session_code (session_id, code)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """))

        # Decisions table
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS decisions (
                id INT PRIMARY KEY AUTO_INCREMENT,
                session_id VARCHAR(36),
                participant_id VARCHAR(36),
                round_number INT,
                choice VARCHAR(1),
                cost DECIMAL(10,2),
                total_cost DECIMAL(10,2),
                payout DECIMAL(10,2),
                others_A INT,
                created_at VARCHAR(30),
                INDEX idx_session_round (session_id, round_number),
                INDEX idx_participant_round (participant_id, round_number),
                UNIQUE KEY ux_participant_round (participant_id, round_number)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """))

        # Migrate existing decisions table - add missing columns if they don't exist
        try:
            conn.execute(text("""
                ALTER TABLE decisions
                ADD COLUMN IF NOT EXISTS total_cost DECIMAL(10,2),
                ADD COLUMN IF NOT EXISTS payout DECIMAL(10,2),
                ADD COLUMN IF NOT EXISTS others_A INT
            """))
        except Exception as e:
            # MySQL doesn't support IF NOT EXISTS in ALTER TABLE, try individual columns
            for col_def in [
                ("total_cost", "DECIMAL(10,2)"),
                ("payout", "DECIMAL(10,2)"),
                ("others_A", "INT")
            ]:
                try:
                    conn.execute(text(f"ALTER TABLE decisions ADD COLUMN {col_def[0]} {col_def[1]}"))
                except Exception:
                    pass  # Column already exists

        conn.commit()

    print("✅ Database schema initialized successfully!")


def background_thread():
    """Example of how to send server generated events to clients."""
    count = 0
    while True:
        socketio.sleep(10)
        count += 1
        socketio.emit('my_response',
                      {'data': 'Server generated event', 'count': count})


@app.route('/')
def index():
    """Redirect to join page as default landing page for participants."""
    return redirect(url_for('join'))


# STEP 2: Test route to verify app still works after changes
@app.route('/healthz')
@admin_required
def healthz():
    return "OK - Step 2: Test route works!", 200


# STEP 3: Test DB connection
@app.route('/test_db')
def test_db():
    try:
        with get_db() as conn:
            result = conn.execute(text("SELECT 1"))
            return "OK - Step 3: DB connection works!", 200
    except Exception as e:
        return f"ERROR - Step 3: {str(e)}", 500


# STEP 5: Initialize database schema
@app.route('/init_db')
@admin_required
def init_db_route():
    """Initialize database schema (admin only)."""
    try:
        init_db()
        return "OK - Step 5: Database schema initialized!", 200
    except Exception as e:
        return f"ERROR - Step 5: {str(e)}", 500


# STEP 5.1: Migrate existing tables (add missing columns)
@app.route('/migrate_db')
@admin_required
def migrate_db():
    """Add missing columns to existing tables."""
    try:
        with get_db() as conn:
            # Check and add 'status' column to sessions table
            try:
                conn.execute(text("SELECT status FROM sessions LIMIT 1"))
            except:
                conn.execute(text("ALTER TABLE sessions ADD COLUMN status VARCHAR(20) DEFAULT 'lobby'"))

            # Check and add 'created_at' column to participants table
            try:
                conn.execute(text("SELECT created_at FROM participants LIMIT 1"))
            except:
                conn.execute(text("ALTER TABLE participants ADD COLUMN created_at VARCHAR(30)"))

            # Check and add 'ready_for_next' column to participants table
            try:
                conn.execute(text("SELECT ready_for_next FROM participants LIMIT 1"))
            except:
                conn.execute(text("ALTER TABLE participants ADD COLUMN ready_for_next TINYINT DEFAULT 0"))

            conn.commit()

        return "OK - Database migrated! Added status, created_at, ready_for_next columns.", 200
    except Exception as e:
        return f"ERROR - Migration failed: {str(e)}", 500


# STEP 4: Admin routes
@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    """Admin login page."""
    if request.method == "POST":
        password = request.form.get("password", "").strip()
        if password == ADMIN_PASSWORD:
            session.permanent = True  # Make session persistent
            session["admin_ok"] = True
            return redirect(url_for("admin_dashboard"))
        else:
            return render_template("admin_login.html", error="Falsches Passwort!")
    return render_template("admin_login.html", error=None)

@app.route("/admin_logout")
def admin_logout():
    """Admin logout."""
    session.pop("admin_ok", None)
    return redirect(url_for("admin_login"))

@app.route("/admin", methods=["GET", "POST"])
@admin_required
def admin_dashboard():
    """Admin dashboard - shows all sessions."""
    with get_db() as conn:
        # POST: Create new session
        if request.method == "POST":
            name = request.form.get("name", f"Session {datetime.datetime.now():%Y-%m-%d %H:%M}")
            group_size = int(request.form.get("group_size", "6"))
            rounds = int(request.form.get("rounds", "20"))
            starting_balance = float(request.form.get("base_payout", "500"))

            # Create session
            session_id = str(uuid.uuid4())
            conn.execute(text("""
                INSERT INTO sessions (id, name, group_size, rounds, starting_balance, created_at, archived, status)
                VALUES (:id, :name, :group_size, :rounds, :starting_balance, :created_at, 0, 'lobby')
            """), {
                "id": session_id,
                "name": name,
                "group_size": group_size,
                "rounds": rounds,
                "starting_balance": starting_balance,
                "created_at": iso_utc(utc_now())
            })

            # Create participants with unique codes
            for i in range(group_size):
                participant_id = str(uuid.uuid4())

                # Generate unique code
                while True:
                    code = create_code(6)
                    existing = conn.execute(text("SELECT 1 FROM participants WHERE code = :code"), {"code": code}).fetchone()
                    if not existing:
                        break

                conn.execute(text("""
                    INSERT INTO participants
                    (id, session_id, code, joined, join_number, current_round, balance, completed, created_at, ready_for_next)
                    VALUES (:id, :session_id, :code, 0, NULL, 1, :balance, 0, :created_at, 0)
                """), {
                    "id": participant_id,
                    "session_id": session_id,
                    "code": code,
                    "balance": starting_balance,
                    "created_at": iso_utc(utc_now())
                })

            conn.commit()

            # Notify all admin clients via Socket.IO
            socketio.emit('session_created', {'session_id': session_id}, room='admin_room')

            return redirect(url_for("admin_dashboard"))

        # GET: Show all sessions
        result = conn.execute(text("""
            SELECT id, name, group_size, rounds, starting_balance, created_at, archived, status
            FROM sessions
            ORDER BY created_at DESC
        """))
        sessions = [dict(row._mapping) for row in result]

        # Get participants for each session
        for s in sessions:
            p_result = conn.execute(text("SELECT code, joined FROM participants WHERE session_id = :sid"), {"sid": s["id"]})
            s["participants"] = [dict(row._mapping) for row in p_result]

    # Separate sessions by status
    sessions_active = [s for s in sessions if not s["archived"] and s["status"] == "lobby"]
    sessions_playing = [s for s in sessions if not s["archived"] and s["status"] == "playing"]
    sessions_done = [s for s in sessions if not s["archived"] and s["status"] == "done"]
    sessions_arch = [s for s in sessions if s["archived"]]

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    return render_template("admin_dashboard.html",
                         sessions_active=sessions_active,
                         sessions_playing=sessions_playing,
                         sessions_done=sessions_done,
                         sessions_arch=sessions_arch,
                         now=now)


@app.route("/admin/delete_session", methods=["POST"])
@admin_required
def admin_delete_session():
    """Delete a session and all its participants."""
    session_id = request.form.get("session_id")

    with get_db() as conn:
        # Delete decisions first (if any)
        conn.execute(text("DELETE FROM decisions WHERE session_id = :sid"), {"sid": session_id})

        # Delete participants
        conn.execute(text("DELETE FROM participants WHERE session_id = :sid"), {"sid": session_id})

        # Delete session
        conn.execute(text("DELETE FROM sessions WHERE id = :sid"), {"sid": session_id})

        conn.commit()

    # Notify admin clients
    socketio.emit('session_deleted', {'session_id': session_id}, room='admin_room')

    return redirect(url_for("admin_dashboard"))


@app.route("/admin/reset_session", methods=["POST"])
@admin_required
def admin_reset_session():
    """Reset a session to lobby state - participants must rejoin with their codes."""
    session_id = request.form.get("session_id")

    with get_db() as conn:
        # Get starting balance
        s_result = conn.execute(text("""
            SELECT starting_balance FROM sessions WHERE id = :sid
        """), {"sid": session_id})
        session_data = s_result.fetchone()

        if not session_data:
            return "Session not found", 404

        starting_balance = session_data[0]

        # Delete all decisions
        conn.execute(text("DELETE FROM decisions WHERE session_id = :sid"), {"sid": session_id})

        # Reset all participants (keep codes, reset joined status)
        conn.execute(text("""
            UPDATE participants
            SET joined = 0,
                join_number = NULL,
                current_round = 1,
                balance = :balance,
                ready_for_next = 0,
                completed = 0
            WHERE session_id = :sid
        """), {"sid": session_id, "balance": starting_balance})

        # Reset session status to lobby
        conn.execute(text("""
            UPDATE sessions
            SET status = 'lobby'
            WHERE id = :sid
        """), {"sid": session_id})

        conn.commit()

    # Emit Socket.IO event to kick all participants to /join
    socketio.emit('session_reset', {
        'message': 'Session wurde zurückgesetzt. Bitte neu joinen!'
    }, room=f"session_{session_id}")

    # Notify admin clients
    socketio.emit('session_reset_admin', {'session_id': session_id}, room='admin_room')

    return redirect(url_for("admin_dashboard"))


@app.route("/admin/archive_session", methods=["POST"])
@admin_required
def admin_archive_session():
    """Archive a session (set archived=1)."""
    session_id = request.form.get("session_id")

    with get_db() as conn:
        conn.execute(text("UPDATE sessions SET archived = 1 WHERE id = :sid"), {"sid": session_id})
        conn.commit()

    # Notify admin clients
    socketio.emit('session_archived', {'session_id': session_id}, room='admin_room')

    return redirect(url_for("admin_dashboard"))


@app.route("/admin/session/<session_id>")
@admin_required
def admin_session_detail(session_id):
    """Show detailed session view with live updates."""
    with get_db() as conn:
        # Load session data
        s_result = conn.execute(text("""
            SELECT id, name, group_size, rounds, starting_balance, status
            FROM sessions WHERE id = :sid
        """), {"sid": session_id})

        session_data = s_result.fetchone()
        if not session_data:
            return "Session not found", 404

        session_dict = dict(session_data._mapping)

        # Get current round (max of all participants)
        r_result = conn.execute(text("""
            SELECT MAX(current_round) as max_round FROM participants WHERE session_id = :sid
        """), {"sid": session_id})
        round_number = r_result.fetchone()[0] or 1

    return render_template("admin_session.html", session=session_dict, round_number=round_number)


@app.route("/admin_session_status")
@admin_required
def admin_session_status():
    """API endpoint for live session status updates."""
    from flask import jsonify

    session_id = request.args.get("session_id")

    if not session_id:
        return jsonify({"error": "session_id required"}), 400

    try:
        with get_db() as conn:
            # Get session info
            s_result = conn.execute(text("""
                SELECT id, name, group_size, rounds, starting_balance, status
                FROM sessions WHERE id = :sid
            """), {"sid": session_id})
            session_data = s_result.fetchone()

            if not session_data:
                return jsonify({"error": "Session not found"}), 404

            session_dict = dict(session_data._mapping)

            # Get current round
            r_result = conn.execute(text("""
                SELECT MAX(current_round) as max_round FROM participants WHERE session_id = :sid
            """), {"sid": session_id})
            current_round = r_result.fetchone()[0] or 1
            session_dict['current_round'] = current_round

            # Get all participants with their status
            p_result = conn.execute(text("""
                SELECT id, code, join_number, current_round, ptype, balance, ready_for_next
                FROM participants WHERE session_id = :sid
                ORDER BY join_number
            """), {"sid": session_id})

            participants = []
            decided_count = 0
            ready_count = 0

            for p in p_result:
                p_dict = dict(p._mapping)

                # Check if participant has made a decision this round
                try:
                    ch_result = conn.execute(text("""
                        SELECT choice FROM decisions
                        WHERE participant_id = :pid AND round_number = :rnd
                    """), {"pid": p_dict['id'], "rnd": current_round})
                    choice_row = ch_result.fetchone()

                    decided = choice_row is not None
                    if decided:
                        decided_count += 1
                        p_dict['choice'] = choice_row[0]
                    else:
                        p_dict['choice'] = None
                except Exception as e:
                    # Decisions table doesn't exist or query failed
                    print(f"Warning: Could not query decisions: {e}")
                    decided = False
                    p_dict['choice'] = None

                p_dict['decided'] = decided
                p_dict['round_display'] = f"{p_dict['current_round']}/{session_dict['rounds']}"

                if p_dict['ready_for_next']:
                    ready_count += 1

                participants.append(p_dict)

            return jsonify({
                "session": session_dict,
                "participants": participants,
                "decided_count": decided_count,
                "ready_count": ready_count
            })
    except Exception as e:
        print(f"Error in admin_session_status: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/admin_export_session_xlsx")
@admin_required
def admin_export_session_xlsx():
    """Export session data as Excel file with multiple sheets."""
    from flask import Response
    import io

    session_id = request.args.get("session_id")
    if not session_id:
        return "session_id required", 400

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return admin_export_session_csv(session_id)

    with get_db() as conn:
        # Get session info
        s_result = conn.execute(text("""
            SELECT id, name, group_size, rounds, starting_balance, created_at
            FROM sessions WHERE id = :sid
        """), {"sid": session_id})
        session_data = s_result.fetchone()

        if not session_data:
            return "Session not found", 404

        session_dict = dict(session_data._mapping)

        # Get all decisions with full details
        decisions_result = conn.execute(text("""
            SELECT d.round_number, p.join_number, p.code, p.ptype, d.choice,
                   d.a_cost, d.b_cost, d.total_cost, d.payout,
                   d.created_at, d.others_A
            FROM decisions d
            JOIN participants p ON d.participant_id = p.id
            WHERE d.session_id = :sid
            ORDER BY d.round_number, p.join_number
        """), {"sid": session_id})

        # Get all participants
        participants_result = conn.execute(text("""
            SELECT join_number, code, ptype, joined
            FROM participants
            WHERE session_id = :sid
            ORDER BY join_number
        """), {"sid": session_id})

        # Create Excel workbook
        wb = openpyxl.Workbook()

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # ============ DECISIONS SHEET ============
        ws_decisions = wb.active
        ws_decisions.title = "Decisions"

        # Headers for Decisions
        decision_headers = ["round", "player_#", "code", "ptype", "choice",
                           "a_cost", "b_cost", "total_cost", "payout", "created_at", "others"]
        ws_decisions.append(decision_headers)

        # Style headers
        for col in range(1, len(decision_headers) + 1):
            cell = ws_decisions.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write decision data
        for row in decisions_result:
            row_data = [
                row[0],  # round_number
                row[1],  # player_#
                row[2],  # code
                row[3],  # ptype
                row[4],  # choice
                float(row[5]) if row[5] is not None else None,  # a_cost
                float(row[6]) if row[6] is not None else None,  # b_cost
                float(row[7]) if row[7] is not None else None,  # total_cost
                float(row[8]) if row[8] is not None else None,  # payout
                str(row[9]) if row[9] else "",  # created_at
                row[10] if row[10] is not None else ""  # others_A
            ]
            ws_decisions.append(row_data)

        # Auto-adjust column widths for Decisions
        for column in ws_decisions.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_decisions.column_dimensions[column_letter].width = adjusted_width

        # ============ PARTICIPANTS SHEET ============
        ws_participants = wb.create_sheet("Participants")

        # Headers for Participants
        participant_headers = ["player_no", "code", "ptype", "joined"]
        ws_participants.append(participant_headers)

        # Style headers
        for col in range(1, len(participant_headers) + 1):
            cell = ws_participants.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write participant data
        for row in participants_result:
            ws_participants.append(list(row))

        # Auto-adjust column widths for Participants
        for column in ws_participants.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_participants.column_dimensions[column_letter].width = adjusted_width

        # ============ SESSION SHEET ============
        ws_session = wb.create_sheet("Session")

        # Headers for Session
        session_headers = ["id", "name", "group_size", "rounds", "starting_balance", "created_at"]
        ws_session.append(session_headers)

        # Style headers
        for col in range(1, len(session_headers) + 1):
            cell = ws_session.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write session data
        session_row = [
            session_dict['id'],
            session_dict['name'],
            session_dict['group_size'],
            session_dict['rounds'],
            float(session_dict['starting_balance']) if session_dict['starting_balance'] else None,
            str(session_dict['created_at']) if session_dict.get('created_at') else ""
        ]
        ws_session.append(session_row)

        # Auto-adjust column widths for Session
        for column in ws_session.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_session.column_dimensions[column_letter].width = adjusted_width

        # ============ TYPE COST TABLE SHEET ============
        ws_type_cost = wb.create_sheet("TypeCostTable")

        # Headers for TypeCostTable
        type_cost_headers = ["Typ", "A_cost", "B_cost_0", "B_cost_1", "B_cost_2", "B_cost_3", "B_cost_4"]
        ws_type_cost.append(type_cost_headers)

        # Style headers
        for col in range(1, len(type_cost_headers) + 1):
            cell = ws_type_cost.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write TypeCost data from TYPE_COST configuration
        for ptype, costs in TYPE_COST.items():
            b_costs = costs["B"] if isinstance(costs["B"], list) else [costs["B"]]
            # Pad with None if less than 5 B_cost values
            while len(b_costs) < 5:
                b_costs.append(None)

            row_data = [
                ptype,
                costs["A"],
                b_costs[0] if len(b_costs) > 0 else None,
                b_costs[1] if len(b_costs) > 1 else None,
                b_costs[2] if len(b_costs) > 2 else None,
                b_costs[3] if len(b_costs) > 3 else None,
                b_costs[4] if len(b_costs) > 4 else None
            ]
            ws_type_cost.append(row_data)

        # Auto-adjust column widths for TypeCostTable
        for column in ws_type_cost.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws_type_cost.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"session_{session_dict['name'].replace(' ', '_')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )


def admin_export_session_csv(session_id):
    """Fallback CSV export if openpyxl not available."""
    from flask import Response
    import io
    import csv

    with get_db() as conn:
        # Get session info
        s_result = conn.execute(text("""
            SELECT name, group_size, rounds, starting_balance, status
            FROM sessions WHERE id = :sid
        """), {"sid": session_id})
        session_data = s_result.fetchone()

        if not session_data:
            return "Session not found", 404

        session_dict = dict(session_data._mapping)

        # Get all decisions
        ch_result = conn.execute(text("""
            SELECT p.code, p.join_number, d.round_number, d.choice, d.total_cost, d.payout
            FROM decisions d
            JOIN participants p ON d.participant_id = p.id
            WHERE d.session_id = :sid
            ORDER BY d.round_number, p.join_number
        """), {"sid": session_id})

        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)

        # Write session info
        writer.writerow(["Session Name:", session_dict['name']])
        writer.writerow(["Group Size (N):", session_dict['group_size']])
        writer.writerow(["Rounds (R):", session_dict['rounds']])
        writer.writerow(["Starting Balance (M):", session_dict['starting_balance']])
        writer.writerow(["Status:", session_dict['status']])
        writer.writerow([])  # Empty row

        # Write header
        writer.writerow(["Code", "Player #", "Round", "Choice", "Cost", "Payout"])

        # Write data
        for row in ch_result:
            # Convert Decimal to float for CSV compatibility
            row_data = [
                row[0],  # code
                row[1],  # join_number
                row[2],  # round_number
                row[3],  # choice
                float(row[4]) if row[4] is not None else None,  # total_cost
                float(row[5]) if row[5] is not None else None   # payout
            ]
            writer.writerow(row_data)

        # Create response
        filename = f"session_{session_dict['name'].replace(' ', '_')}.csv"
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )


# STEP 7: Participant routes
@app.route("/join", methods=["GET", "POST"])
def join():
    """Join a session with participant code."""
    if request.method == "POST":
        code = request.form.get("code", "").strip().upper()

        with get_db() as conn:
            # Find participant by code
            result = conn.execute(text("""
                SELECT p.id, p.session_id, p.joined, p.join_number, s.status, s.archived
                FROM participants p
                JOIN sessions s ON p.session_id = s.id
                WHERE p.code = :code
            """), {"code": code})

            participant = result.fetchone()

            if not participant:
                return render_template("join.html", error="Ungültiger Code!")

            participant = dict(participant._mapping)

            if participant["archived"]:
                return render_template("join.html", error="Diese Session ist archiviert!")

            current_participant_id = session.get("participant_id")

            if current_participant_id and current_participant_id != participant["id"]:
                session.clear()

            # Check for duplicate login (prevent same code in multiple tabs/browsers)
            # Get browser_token from form (sent by JavaScript from localStorage)
            browser_token = request.form.get("browser_token", "").strip()
            if not browser_token:
                # Fallback: generate new token (shouldn't happen with JavaScript)
                browser_token = str(uuid.uuid4())

            participant_id = participant["id"]

            with active_participants_lock:
                if participant_id in active_participants:
                    existing = active_participants[participant_id]
                    time_since_activity = time.time() - existing['last_activity']

                    # If active within last 60 seconds AND different browser token
                    if time_since_activity < 60 and existing['browser_token'] != browser_token:
                        print(f"⚠️ Duplicate login blocked: Code '{code}', existing token: {existing['browser_token'][:20]}..., new token: {browser_token[:20]}...")
                        return render_template("join.html",
                            error="Dieser Code wird bereits verwendet! Bitte warten Sie 60 Sekunden oder schließen Sie den anderen Tab.")

                # Register/update this participant as active
                active_participants[participant_id] = {
                    'browser_token': browser_token,
                    'last_activity': time.time()
                }
                print(f"✅ Login registered: Code '{code}', token: {browser_token[:20]}...")

            session["participant_id"] = participant["id"]
            session["session_id"] = participant["session_id"]
            session["code"] = code
            session["browser_token"] = browser_token  # Store for validation

            if participant["joined"]:
                p_full = conn.execute(text("""
                    SELECT current_round, ptype
                    FROM participants
                    WHERE id = :pid
                """), {"pid": participant["id"]}).fetchone()

                if participant["status"] == "lobby":
                    return redirect(url_for("lobby"))
                elif participant["status"] == "playing":
                    r = p_full[0] or 1
                    decided_check = conn.execute(text("""
                        SELECT 1 FROM decisions
                        WHERE participant_id = :pid AND round_number = :r
                    """), {"pid": participant["id"], "r": r}).fetchone()

                    if decided_check:
                        return redirect(url_for("wait_view"))
                    else:
                        return redirect(url_for("round_view"))
                elif participant["status"] == "done":
                    return redirect(url_for("done"))
                else:
                    return redirect(url_for("lobby"))

            if participant["status"] != "lobby":
                return render_template("join.html", error="Diese Session ist bereits gestartet oder beendet!")

            if not participant["joined"]:
                # Get max join_number for this session
                max_result = conn.execute(text("""
                    SELECT COALESCE(MAX(join_number), 0) as max_num
                    FROM participants
                    WHERE session_id = :sid AND joined = 1
                """), {"sid": participant["session_id"]})
                max_num = max_result.fetchone()[0]

                # Assign next join_number
                new_join_number = max_num + 1

                conn.execute(text("""
                    UPDATE participants
                    SET joined = 1, join_number = :join_num
                    WHERE id = :pid
                """), {"join_num": new_join_number, "pid": participant["id"]})

                conn.commit()

                # Notify all participants in this session's lobby via Socket.IO
                socketio.emit('lobby_update', {
                    'joined': new_join_number
                }, room=f"session_{participant['session_id']}")

                # Also notify admin room
                socketio.emit('lobby_update', {
                    'session_id': participant['session_id'],
                    'joined': new_join_number
                }, room='admin_room')

            return redirect(url_for("lobby"))

    return render_template("join.html", error=None)


@app.route("/lobby")
def lobby():
    """Lobby - wait for all participants to join."""
    participant, session_id = get_participant_state()

    if not participant:
        return redirect(url_for("join"))

    if not participant["joined"]:
        return redirect(url_for("join"))

    if participant["status"] == "playing":
        return redirect(url_for("round_view"))

    if participant["status"] == "done":
        return redirect(url_for("done"))

    with get_db() as conn:
        s_result = conn.execute(text("""
            SELECT id, name, group_size, rounds, starting_balance, status
            FROM sessions
            WHERE id = :sid
        """), {"sid": session_id})

        session_data = s_result.fetchone()

        if not session_data:
            return redirect(url_for("join"))

        session_data = dict(session_data._mapping)

        count_result = conn.execute(text("""
            SELECT COUNT(*) as joined_count
            FROM participants
            WHERE session_id = :sid AND joined = 1
        """), {"sid": session_id})

        joined_count = count_result.fetchone()[0]

        if joined_count >= session_data["group_size"] and session_data["status"] == "lobby":
            conn.execute(text("""
                UPDATE sessions
                SET status = 'playing'
                WHERE id = :sid
            """), {"sid": session_id})

            conn.execute(text("""
                UPDATE participants
                SET current_round = 1
                WHERE session_id = :sid
            """), {"sid": session_id})

            conn.commit()

            return redirect(url_for("round_view"))

    return render_template("lobby.html",
                         session=session_data,
                         participant=participant,
                         joined=joined_count)


@app.route("/round")
def round_view():
    """Display current round with decision options."""
    participant, session_id = get_participant_state()

    if not participant:
        return redirect(url_for("join"))

    if not participant["joined"]:
        return redirect(url_for("join"))

    if participant["status"] == "lobby":
        return redirect(url_for("lobby"))

    if participant["status"] == "done":
        return redirect(url_for("done"))

    r = participant["current_round"] or 1

    if r > participant["rounds"]:
        return redirect(url_for("done"))

    with get_db() as conn:
        already_decided = conn.execute(text("""
            SELECT 1 FROM decisions
            WHERE participant_id = :pid AND round_number = :r
        """), {"pid": participant["id"], "r": r}).fetchone()

        if already_decided:
            return redirect(url_for("wait_view"))

        ptype = participant["ptype"] or 1
        N = participant["group_size"]

        a_cost_display = a_cost_for(ptype)
        others_max = max(1, N - 1)
        b_row_costs = [int(b_cost_adapt(ptype, k, N)) for k in range(1, others_max + 1)]
        b_list = [{"others": k, "cost": b_row_costs[k-1]} for k in range(1, others_max + 1)]

    session_data = {
        "id": session_id,
        "group_size": participant["group_size"],
        "rounds": participant["rounds"],
        "status": participant["status"]
    }

    return render_template(
        "round.html",
        session=session_data,
        round_number=r,
        N=N,
        a_cost_display=a_cost_display,
        b_list=b_list,
        others_max=others_max,
        base_payout=500,
        balance_current=500,
        participant=participant
    )


@app.route("/choose", methods=["POST"])
def choose():
    """Record participant's choice for current round."""
    participant_id = session.get("participant_id")
    session_id = session.get("session_id")

    if not participant_id or not session_id:
        return ("No participant", 400)

    data = request.get_json() or {}
    choice = (data.get("choice") or "").upper()

    if choice not in ("A", "B"):
        return ("Invalid choice", 400)

    with get_db() as conn:
        # Get participant's current round
        p_result = conn.execute(text("""
            SELECT current_round
            FROM participants
            WHERE id = :pid
        """), {"pid": participant_id})

        participant = p_result.fetchone()

        if not participant:
            return ("Participant not found", 404)

        r = participant[0] or 1

        # Check if already decided
        already = conn.execute(text("""
            SELECT 1 FROM decisions
            WHERE participant_id = :pid AND round_number = :r
        """), {"pid": participant_id, "r": r}).fetchone()

        if already:
            return ({"ok": True}, 200)

        # Insert decision
        conn.execute(text("""
            INSERT INTO decisions (session_id, participant_id, round_number, choice, created_at)
            VALUES (:sid, :pid, :r, :choice, :created_at)
        """), {
            "sid": session_id,
            "pid": participant_id,
            "r": r,
            "choice": choice,
            "created_at": iso_utc(utc_now())
        })

        conn.commit()

        # Notify all participants in this round via Socket.IO
        socketio.emit('round_decision', {
            'round': r,
            'decided': True
        }, room=f"round_{session_id}_{r}")

        # Also notify admin room for live updates in detail view
        socketio.emit('round_decision', {
            'session_id': session_id,
            'round': r
        }, room='admin_room')

    return ({"ok": True}, 200)


@app.route("/wait")
def wait_view():
    """Wait for all participants to decide."""
    participant, session_id = get_participant_state()

    if not participant:
        return redirect(url_for("join"))

    if not participant["joined"]:
        return redirect(url_for("join"))

    if participant["status"] == "lobby":
        return redirect(url_for("lobby"))

    if participant["status"] == "done":
        return redirect(url_for("done"))

    r = participant["current_round"] or 1

    if r > participant["rounds"]:
        return redirect(url_for("done"))

    with get_db() as conn:
        already_decided = conn.execute(text("""
            SELECT 1 FROM decisions
            WHERE participant_id = :pid AND round_number = :r
        """), {"pid": participant["id"], "r": r}).fetchone()

        if not already_decided:
            return redirect(url_for("round_view"))

        decided_result = conn.execute(text("""
            SELECT COUNT(*) as c
            FROM decisions
            WHERE session_id = :sid AND round_number = :r
        """), {"sid": session_id, "r": r})

        decided = decided_result.fetchone()[0]

        if decided >= participant["group_size"]:
            return redirect(url_for("reveal"))

    session_data = {
        "id": session_id,
        "group_size": participant["group_size"],
        "rounds": participant["rounds"],
        "status": participant["status"]
    }

    return render_template("wait.html",
                         session=session_data,
                         round_number=r,
                         decided=decided,
                         participant=participant)


def finalize_round(session_id, round_number):
    """Calculate costs and payouts for completed round."""
    with get_db() as conn:
        s_result = conn.execute(text("""
            SELECT group_size, starting_balance
            FROM sessions
            WHERE id = :sid
        """), {"sid": session_id})

        session_data = s_result.fetchone()
        if not session_data:
            return

        session_dict = dict(session_data._mapping)

        decided_result = conn.execute(text("""
            SELECT COUNT(*) as c
            FROM decisions
            WHERE session_id = :sid AND round_number = :r
        """), {"sid": session_id, "r": round_number})

        decided_count = decided_result.fetchone()[0]

        if decided_count < session_dict["group_size"]:
            return

        missing_result = conn.execute(text("""
            SELECT COUNT(*) as c
            FROM decisions
            WHERE session_id = :sid AND round_number = :r AND total_cost IS NULL
        """), {"sid": session_id, "r": round_number})

        missing_count = missing_result.fetchone()[0]

        if missing_count <= 0:
            return

        rows_result = conn.execute(text("""
            SELECT d.id, d.participant_id, d.choice, p.ptype, p.join_number
            FROM decisions d
            JOIN participants p ON p.id = d.participant_id
            WHERE d.session_id = :sid AND d.round_number = :r
            ORDER BY p.join_number
        """), {"sid": session_id, "r": round_number})

        rows = rows_result.fetchall()

        total_A = sum(1 for row in rows if row[2] == "A")
        N = session_dict["group_size"]
        M = float(session_dict["starting_balance"] or 500)

        for row in rows:
            did = row[0]
            pid = row[1]
            choice = row[2]
            ptype = row[3] or 1

            if choice == "A":
                cost = a_cost_for(ptype)
                others_A = max(0, total_A - 1)
            else:
                others_A = total_A
                cost = b_cost_adapt(ptype, others_A, N)

            payout = max(M - float(cost), 0)

            conn.execute(text("""
                UPDATE decisions
                SET total_cost = :cost, payout = :payout, others_A = :others_A
                WHERE id = :did AND total_cost IS NULL
            """), {
                "cost": cost,
                "payout": payout,
                "others_A": others_A,
                "did": did
            })

            conn.execute(text("""
                UPDATE participants
                SET balance = :payout
                WHERE id = :pid
            """), {"payout": payout, "pid": pid})

        conn.commit()


@app.route("/reveal")
def reveal():
    """Show round results."""
    participant, session_id = get_participant_state()

    if not participant:
        return redirect(url_for("join"))

    if not participant["joined"]:
        return redirect(url_for("join"))

    if participant["status"] == "lobby":
        return redirect(url_for("lobby"))

    if participant["status"] == "done":
        return redirect(url_for("done"))

    r = participant["current_round"] or 1

    if r > participant["rounds"]:
        return redirect(url_for("done"))

    with get_db() as conn:
        decided_result = conn.execute(text("""
            SELECT COUNT(*) as c
            FROM decisions
            WHERE session_id = :sid AND round_number = :r
        """), {"sid": session_id, "r": r})

        decided = decided_result.fetchone()[0]

        if decided < participant["group_size"]:
            return redirect(url_for("wait_view"))

        finalize_round(session_id, r)

        is_last_round = (r >= participant["rounds"])

    session_data = {
        "id": session_id,
        "group_size": participant["group_size"],
        "rounds": participant["rounds"],
        "status": participant["status"]
    }

    return render_template("reveal.html",
                         session=session_data,
                         round_number=r,
                         participant=participant,
                         is_last_round=is_last_round)


@app.route("/confirm_ready", methods=["POST"])
def confirm_ready():
    """Player confirms ready for next round."""
    participant_id = session.get("participant_id")
    session_id = session.get("session_id")

    if not participant_id or not session_id:
        return ("No participant", 400)

    with get_db() as conn:
        p_result = conn.execute(text("""
            SELECT current_round
            FROM participants
            WHERE id = :pid
        """), {"pid": participant_id})

        participant = p_result.fetchone()

        if not participant:
            return ("Participant not found", 404)

        current_round = participant[0] or 1

        conn.execute(text("""
            UPDATE participants
            SET ready_for_next = 1
            WHERE id = :pid
        """), {"pid": participant_id})

        conn.commit()

        socketio.emit('player_ready', {}, room=f"session_{session_id}")

        # Also notify admin room
        socketio.emit('player_ready', {'session_id': session_id}, room='admin_room')

        ready_count_result = conn.execute(text("""
            SELECT COUNT(*) as c
            FROM participants
            WHERE session_id = :sid AND ready_for_next = 1
        """), {"sid": session_id})

        ready_count = ready_count_result.fetchone()[0]

        group_size_result = conn.execute(text("""
            SELECT group_size
            FROM sessions
            WHERE id = :sid
        """), {"sid": session_id})

        group_size = group_size_result.fetchone()[0]

        if ready_count >= group_size:
            s_result = conn.execute(text("""
                SELECT rounds
                FROM sessions
                WHERE id = :sid
            """), {"sid": session_id})

            total_rounds = s_result.fetchone()[0]

            if current_round >= total_rounds:
                conn.execute(text("""
                    UPDATE sessions
                    SET status = 'done'
                    WHERE id = :sid
                """), {"sid": session_id})

                conn.execute(text("""
                    UPDATE participants
                    SET ready_for_next = 0
                    WHERE session_id = :sid
                """), {"sid": session_id})

                conn.commit()

                socketio.emit('game_finished', {}, room=f"session_{session_id}")
                socketio.emit('game_finished', {'session_id': session_id}, room='admin_room')
            else:
                conn.execute(text("""
                    UPDATE participants
                    SET current_round = current_round + 1, ready_for_next = 0
                    WHERE session_id = :sid
                """), {"sid": session_id})

                conn.commit()

                socketio.emit('all_ready', {
                    'next_round': current_round + 1
                }, room=f"session_{session_id}")

                # Also notify admin room
                socketio.emit('all_ready', {
                    'session_id': session_id,
                    'next_round': current_round + 1
                }, room='admin_room')

    return ({"ok": True}, 200)


@app.route("/done")
def done():
    """Game finished - show final results."""
    participant, session_id = get_participant_state()

    if not participant:
        return redirect(url_for("join"))

    if not participant["joined"]:
        return redirect(url_for("join"))

    if participant["status"] == "lobby":
        return redirect(url_for("lobby"))

    if participant["status"] == "playing":
        return redirect(url_for("round_view"))

    return render_template("done.html",
                         participant=participant,
                         session_id=session_id)


# API: Check current session for tab-isolation
@app.route("/api/check_session")
def api_check_session():
    """Return current participant code from session (for tab-isolation)."""
    from flask import jsonify

    code = session.get("code")
    participant_id = session.get("participant_id")

    return jsonify({
        "code": code,
        "participant_id": participant_id,
        "has_session": bool(participant_id)
    })


@socketio.event
def my_event(message):
    session['receive_count'] = session.get('receive_count', 0) + 1
    emit('my_response',
         {'data': message['data'], 'count': session['receive_count']})


@socketio.event
def my_broadcast_event(message):
    session['receive_count'] = session.get('receive_count', 0) + 1
    emit('my_response',
         {'data': message['data'], 'count': session['receive_count']},
         broadcast=True)


@socketio.event
def join(message):
    join_room(message['room'])
    session['receive_count'] = session.get('receive_count', 0) + 1
    emit('my_response',
         {'data': 'In rooms: ' + ', '.join(rooms()),
          'count': session['receive_count']})


@socketio.event
def leave(message):
    leave_room(message['room'])
    session['receive_count'] = session.get('receive_count', 0) + 1
    emit('my_response',
         {'data': 'In rooms: ' + ', '.join(rooms()),
          'count': session['receive_count']})


@socketio.on('close_room')
def on_close_room(message):
    session['receive_count'] = session.get('receive_count', 0) + 1
    emit('my_response', {'data': 'Room ' + message['room'] + ' is closing.',
                         'count': session['receive_count']},
         to=message['room'])
    close_room(message['room'])


@socketio.event
def my_room_event(message):
    session['receive_count'] = session.get('receive_count', 0) + 1
    emit('my_response',
         {'data': message['data'], 'count': session['receive_count']},
         to=message['room'])


@socketio.on('*')
def catch_all(event, data):
    session['receive_count'] = session.get('receive_count', 0) + 1
    emit('my_response',
         {'data': [event, data], 'count': session['receive_count']})


@socketio.event
def disconnect_request():
    @copy_current_request_context
    def can_disconnect():
        disconnect()

    session['receive_count'] = session.get('receive_count', 0) + 1
    # for this emit we use a callback function
    # when the callback function is invoked we know that the message has been
    # received and it is safe to disconnect
    emit('my_response',
         {'data': 'Disconnected!', 'count': session['receive_count']},
         callback=can_disconnect)


@socketio.event
def my_ping():
    emit('my_pong')


@socketio.event
def connect():
    global thread
    with thread_lock:
        if thread is None:
            thread = socketio.start_background_task(background_thread)
    emit('my_response', {'data': 'Connected', 'count': 0})


# STEP 4: Admin Socket.IO events
@socketio.on('admin_connect')
def handle_admin_connect():
    """Admin connects to dashboard - join admin room for broadcasts."""
    if not require_admin():
        return {'error': 'Unauthorized'}

    join_room('admin_room')
    emit('admin_status', {'message': 'Connected to admin live updates'})
    print(f"Admin connected via Socket.IO: {request.sid}")

@socketio.on('admin_get_sessions')
def handle_admin_get_sessions():
    """Get all sessions from database."""
    if not require_admin():
        return {'error': 'Unauthorized'}

    with get_db() as conn:
        result = conn.execute(text("""
            SELECT id, name, group_size, rounds, starting_balance, created_at, archived, status
            FROM sessions
            ORDER BY created_at DESC
        """))
        sessions = []
        for row in result:
            s = dict(row._mapping)
            # Get participants for this session
            p_result = conn.execute(text("SELECT code, joined FROM participants WHERE session_id = :sid"), {"sid": s["id"]})
            s["participants"] = [dict(p._mapping) for p in p_result]
            sessions.append(s)

    emit('admin_sessions_update', {'sessions': sessions})


# STEP 7: Participant Socket.IO events
@socketio.on('join_lobby')
def handle_join_lobby(data):
    """Participant joins lobby room for real-time updates."""
    session_id = session.get("session_id")
    participant_id = session.get("participant_id")

    if not session_id or not participant_id:
        return {'error': 'Not authenticated'}

    # Join the session-specific room
    room = f"session_{session_id}"
    join_room(room)

    # Send current lobby status to this participant
    with get_db() as conn:
        # Count joined participants
        count_result = conn.execute(text("""
            SELECT COUNT(*) as joined_count
            FROM participants
            WHERE session_id = :sid AND joined = 1
        """), {"sid": session_id})

        joined_count = count_result.fetchone()[0]

        # Get session info
        s_result = conn.execute(text("""
            SELECT group_size, status
            FROM sessions
            WHERE id = :sid
        """), {"sid": session_id})

        session_data = s_result.fetchone()

        if session_data:
            session_dict = dict(session_data._mapping)
            emit('lobby_status', {
                'joined': joined_count,
                'group_size': session_dict['group_size'],
                'ready': joined_count >= session_dict['group_size']
            })

    print(f"Participant {participant_id} joined lobby room: {room}")


@socketio.on('join_round')
def handle_join_round(data):
    """Participant joins round room for real-time decision updates."""
    session_id = session.get("session_id")
    participant_id = session.get("participant_id")

    if not session_id or not participant_id:
        return {'error': 'Not authenticated'}

    with get_db() as conn:
        # Get participant's current round
        p_result = conn.execute(text("""
            SELECT current_round
            FROM participants
            WHERE id = :pid
        """), {"pid": participant_id})

        participant = p_result.fetchone()

        if not participant:
            return {'error': 'Participant not found'}

        r = participant[0] or 1

        # Join the round-specific room
        room = f"round_{session_id}_{r}"
        join_room(room)

        # Send current round status to this participant
        # Count decisions for this round
        decided_result = conn.execute(text("""
            SELECT COUNT(*) as c
            FROM decisions
            WHERE session_id = :sid AND round_number = :r
        """), {"sid": session_id, "r": r})

        decided_count = decided_result.fetchone()[0]

        # Get session group size
        s_result = conn.execute(text("""
            SELECT group_size
            FROM sessions
            WHERE id = :sid
        """), {"sid": session_id})

        session_data = s_result.fetchone()

        if session_data:
            group_size = session_data[0]

            # Get list of players who decided
            decided_players_result = conn.execute(text("""
                SELECT p.join_number
                FROM decisions d
                JOIN participants p ON p.id = d.participant_id
                WHERE d.session_id = :sid AND d.round_number = :r
                ORDER BY p.join_number
            """), {"sid": session_id, "r": r})

            decided_players = [row[0] for row in decided_players_result.fetchall()]

            emit('round_status', {
                'decided': decided_count,
                'group_size': group_size,
                'decided_players': decided_players,
                'ready': decided_count >= group_size
            })

    print(f"Participant {participant_id} joined round room: {room}")


@socketio.on('join_reveal')
def handle_join_reveal(data):
    """Participant joins reveal room for ready status updates."""
    session_id = session.get("session_id")
    participant_id = session.get("participant_id")

    if not session_id or not participant_id:
        return {'error': 'Not authenticated'}

    with get_db() as conn:
        p_result = conn.execute(text("""
            SELECT current_round
            FROM participants
            WHERE id = :pid
        """), {"pid": participant_id})

        participant = p_result.fetchone()

        if not participant:
            return {'error': 'Participant not found'}

        r = participant[0] or 1

        room = f"session_{session_id}"
        join_room(room)

        print(f"[join_reveal] Participant {participant_id} joining reveal for session {session_id}, round {r}")

        try:
            results_result = conn.execute(text("""
                SELECT p.join_number, d.choice, d.total_cost, d.payout
                FROM decisions d
                JOIN participants p ON p.id = d.participant_id
                WHERE d.session_id = :sid AND d.round_number = :r
                ORDER BY p.join_number
            """), {"sid": session_id, "r": r})

            players = []
            for row in results_result.fetchall():
                players.append({
                    "player_no": row[0],
                    "choice": row[1],
                    "cost": float(row[2]) if row[2] is not None else 0,
                    "payout": float(row[3]) if row[3] is not None else 0
                })

            print(f"[join_reveal] Found {len(players)} players with decisions")
        except Exception as e:
            print(f"[join_reveal] ERROR fetching decisions: {e}")
            players = []

        ready_result = conn.execute(text("""
            SELECT p.join_number, p.ready_for_next, p.id
            FROM participants p
            WHERE p.session_id = :sid
            ORDER BY p.join_number
        """), {"sid": session_id})

        ready_players = []
        me_ready = False
        for row in ready_result.fetchall():
            is_ready = bool(row[1])
            ready_players.append({
                "player_no": row[0],
                "ready": is_ready
            })
            if row[2] == participant_id:
                me_ready = is_ready

        ready_count = sum(1 for p in ready_players if p["ready"])

        s_result = conn.execute(text("""
            SELECT group_size
            FROM sessions
            WHERE id = :sid
        """), {"sid": session_id})

        group_size = s_result.fetchone()[0]

        emit('reveal_status', {
            'players': players,
            'ready_count': ready_count,
            'group_size': group_size,
            'all_ready': ready_count >= group_size,
            'me_ready': me_ready,
            'ready_players': ready_players
        })

    print(f"Participant {participant_id} joined reveal room: {room}")


@socketio.on('ready_update')
def handle_ready_update(data):
    """Broadcast when a player marks themselves as ready."""
    session_id = session.get("session_id")

    if not session_id:
        return {'error': 'Not authenticated'}

    socketio.emit('player_ready', {}, room=f"session_{session_id}")


@socketio.on('disconnect')
def test_disconnect(reason):
    print('Client disconnected', request.sid, reason)


if __name__ == '__main__':
    socketio.run(app)
