import os, uuid, random, string, datetime, io
from datetime import timedelta, timezone
from functools import wraps
from flask import (
    Flask, request, redirect, render_template, session as flask_session,
    url_for, jsonify, g, send_file, has_app_context
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pymysql
from pymysql.cursors import DictCursor
from contextlib import contextmanager

APP_DIR = os.path.dirname(os.path.abspath(__file__))

def must_get_env(name: str) -> str:
    """Read required environment variables (fail fast if missing)."""
    val = os.environ.get(name)
    if not val or not val.strip():
        raise RuntimeError(f"Missing required environment variable: {name}")
    return val

ADMIN_PASSWORD = must_get_env("ADMIN_PASSWORD")

# MySQL Configuration
DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_USER = must_get_env("DB_USER")
DB_PASSWORD = must_get_env("DB_PASSWORD")
DB_NAME = must_get_env("DB_NAME")
DB_PORT = int(os.environ.get("DB_PORT", "3306"))

app = Flask(
    __name__,
    template_folder=os.path.join(APP_DIR, "templates"),
    static_folder=os.path.join(APP_DIR, "static"),
)

app.secret_key = must_get_env("SECRET_KEY")

app.config["SESSION_PERMANENT"] = False

DEBUG_MODE = os.environ.get("FLASK_DEBUG", "0") == "1"
app.config["TEMPLATES_AUTO_RELOAD"] = DEBUG_MODE


# -------------------- Vaccination: Cost types --------------------
TYPE_COST = {
    1: {"B": [4, 3, 2, 1, 0],  "A": 4},
    2: {"B": [8, 6, 4, 2, 0],  "A": 4},
    3: {"B": [4, 3, 2, 1, 0],  "A": 8},
    4: {"B": [8, 6, 4, 2, 0],  "A": 8},
    5: {"B": [24, 18, 12, 6, 0], "A": 32},
    6: {"B": [64, 48, 32, 16, 0], "A": 32},
}
B_COLS = 5

def a_cost_for(ptype: int) -> float:
    return TYPE_COST.get(ptype, TYPE_COST[1])["A"]

def b_cost_adapt(ptype: int, others_A: int, N: int) -> float:
    if ptype not in TYPE_COST:
        ptype = 1
    b = TYPE_COST[ptype]["B"]
    N = max(1, int(N))
    others_A = max(0, min(int(others_A), max(0, N-1)))
    if N <= 1:
        return float(b[0])
    frac = others_A / float(N - 1)
    x = frac * B_COLS
    col = int(x + 0.5)
    col = max(1, min(B_COLS, col))
    return float(b[col - 1])

class MySQLConnectionWrapper:

    def __init__(self, conn):
        self._conn = conn

    def execute(self, query, params=None):

        cursor = self._conn.cursor()
        cursor.execute(query, params or ())
        return cursor

    def commit(self):
        return self._conn.commit()

    def rollback(self):
        return self._conn.rollback()

    def close(self):
        return self._conn.close()

    def cursor(self):
        return self._conn.cursor()
    
    def begin(self):
        cursor = self._conn.cursor()
        cursor.execute("START TRANSACTION")
        cursor.close()

 

def _connect_mysql():
    conn = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        port=DB_PORT,
        cursorclass=DictCursor,
        charset='utf8mb4',
        autocommit=False,
        connect_timeout=10,
        read_timeout=30,
        write_timeout=30
    )

    return MySQLConnectionWrapper(conn)
# -------------------- DB helpers --------------------


def db():

    if not has_app_context():
        return _connect_mysql()

    if "db" not in g:
        g.db = _connect_mysql()
    return g.db

@app.teardown_appcontext
def close_db(exception=None):
    con = g.pop("db", None)
    if con is not None:
        try:
            con.close()
        except Exception:
            pass


def ensure_column(con, table, column, definition):
    cursor = con.cursor()
    cursor.execute(f"SHOW COLUMNS FROM {table} LIKE %s", (column,))
    if not cursor.fetchone():
        cursor.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
        con.commit()
    cursor.close()

def ensure_archive_schema(con, base_table):
    arch_table = f"archived_{base_table}"
    cursor = con.cursor()

    # Get base table columns
    cursor.execute(f"SHOW COLUMNS FROM {base_table}")
    base_cols = {row['Field']: row for row in cursor.fetchall()}

    # Get archive table columns
    cursor.execute(f"SHOW COLUMNS FROM {arch_table}")
    arch_cols = {row['Field'] for row in cursor.fetchall()}

    # Add missing columns
    for name, col_info in base_cols.items():
        if name not in arch_cols:
            col_type = col_info['Type']
            default = col_info['Default']
            null = "NULL" if col_info['Null'] == 'YES' else "NOT NULL"

            if default is not None:
                cursor.execute(f"ALTER TABLE {arch_table} ADD COLUMN {name} {col_type} {null} DEFAULT {default}")
            else:
                cursor.execute(f"ALTER TABLE {arch_table} ADD COLUMN {name} {col_type} {null}")

    con.commit()
    cursor.close()

# ---------- UTC helpers (aware) ----------
def utc_now():
    return datetime.datetime.now(timezone.utc).replace(microsecond=0)

def iso_utc(dt: datetime.datetime) -> str:
    return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

def parse_iso_utc(s: str) -> datetime.datetime:
    return datetime.datetime.fromisoformat((s or "").replace("Z", "+00:00"))


def init_db():
    con = db()
    cursor = con.cursor()

    cursor.execute(
        """CREATE TABLE IF NOT EXISTS sessions (
        id VARCHAR(36) PRIMARY KEY,
        name VARCHAR(255),
        group_size INT,
        rounds INT,
        cvac DECIMAL(10,2),
        alpha DECIMAL(10,2),
        cinf DECIMAL(10,2),
        subsidy TINYINT DEFAULT 0,
        subsidy_amount DECIMAL(10,2) DEFAULT 0,
        regime VARCHAR(50),
        starting_balance DECIMAL(10,2) DEFAULT 500,
        created_at VARCHAR(30),
        archived TINYINT DEFAULT 0,
        reveal_window INT DEFAULT 5,
        watch_time INT DEFAULT 15,
        cost_mode VARCHAR(50) DEFAULT 'type_table'
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"""
    )

    cursor.execute(
        """CREATE TABLE IF NOT EXISTS participants (
        id VARCHAR(36) PRIMARY KEY,
        session_id VARCHAR(36),
        code VARCHAR(10) UNIQUE,
        theta DECIMAL(10,2),
        lambda DECIMAL(10,2),
        joined TINYINT DEFAULT 0,
        join_number INT,
        current_round INT DEFAULT 1,
        balance DECIMAL(10,2) DEFAULT 0,
        completed TINYINT DEFAULT 0,
        created_at VARCHAR(30),
        ptype INT,
        ready_for_next TINYINT DEFAULT 0,
        INDEX idx_session (session_id),
        INDEX idx_session_code (session_id, code)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"""
    )

    cursor.execute(
        """CREATE TABLE IF NOT EXISTS decisions (
        id INT PRIMARY KEY AUTO_INCREMENT,
        session_id VARCHAR(36),
        participant_id VARCHAR(36),
        round_number INT,
        choice VARCHAR(1),
        a_cost DECIMAL(10,2),
        b_cost DECIMAL(10,2),
        total_cost DECIMAL(10,2),
        created_at VARCHAR(30),
        reveal TINYINT,
        payout DECIMAL(10,2),
        others_A INT,
        b_cost_round DECIMAL(10,2),
        base_payout DECIMAL(10,2),
        INDEX idx_session_round (session_id, round_number),
        INDEX idx_participant_round (participant_id, round_number),
        UNIQUE KEY ux_participant_round (participant_id, round_number)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"""
    )

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS round_phases (
            session_id VARCHAR(36),
            round_number INT,
            decision_ends_at VARCHAR(30),
            watch_ends_at VARCHAR(30),
            created_at VARCHAR(30),
            PRIMARY KEY (session_id, round_number)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)

    # Create archived tables with same structure
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS archived_sessions (
            id VARCHAR(36) PRIMARY KEY,
            name VARCHAR(255),
            group_size INT,
            rounds INT,
            cvac DECIMAL(10,2),
            alpha DECIMAL(10,2),
            cinf DECIMAL(10,2),
            subsidy TINYINT DEFAULT 0,
            subsidy_amount DECIMAL(10,2) DEFAULT 0,
            regime VARCHAR(50),
            starting_balance DECIMAL(10,2) DEFAULT 500,
            created_at VARCHAR(30),
            archived TINYINT DEFAULT 0,
            reveal_window INT DEFAULT 5,
            watch_time INT DEFAULT 15,
            cost_mode VARCHAR(50) DEFAULT 'type_table'
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS archived_participants (
            id VARCHAR(36) PRIMARY KEY,
            session_id VARCHAR(36),
            code VARCHAR(10),
            theta DECIMAL(10,2),
            lambda DECIMAL(10,2),
            joined TINYINT DEFAULT 0,
            join_number INT,
            current_round INT DEFAULT 1,
            balance DECIMAL(10,2) DEFAULT 0,
            completed TINYINT DEFAULT 0,
            created_at VARCHAR(30),
            ptype INT,
            ready_for_next TINYINT DEFAULT 0
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS archived_decisions (
            id INT PRIMARY KEY,
            session_id VARCHAR(36),
            participant_id VARCHAR(36),
            round_number INT,
            choice VARCHAR(1),
            a_cost DECIMAL(10,2),
            b_cost DECIMAL(10,2),
            total_cost DECIMAL(10,2),
            created_at VARCHAR(30),
            reveal TINYINT,
            payout DECIMAL(10,2),
            others_A INT,
            b_cost_round DECIMAL(10,2),
            base_payout DECIMAL(10,2)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)

    con.commit()
    cursor.close()
    con.close()


# -------------------- Context --------------------
@app.before_request
def load_participant():
    pid = flask_session.get("participant_id")
    g.participant = None
    if pid:
        con = db()
        g.participant = con.execute("SELECT * FROM participants WHERE id=%s", (pid,)).fetchone()

def create_code(n=6):
    chars = (string.ascii_uppercase + string.digits).replace("O","").replace("0","").replace("I","").replace("1","")
    return "".join(random.choice(chars) for _ in range(n))


# -------------------- State & Guard --------------------
def current_state(con, p, s) -> str:
    if not p or not s: return "lobby"
    if s["archived"]: return "done"

    joined = con.execute(
        "SELECT COUNT(*) c FROM participants WHERE session_id=%s AND joined=1", (s["id"],)
    ).fetchone()["c"]
    if joined < s["group_size"]:
        return "lobby"

    r = p["current_round"]

    if r > s["rounds"]:
        all_ready = con.execute(
            "SELECT COUNT(*) c FROM participants WHERE session_id=%s AND ready_for_next=1",
            (s["id"],)
        ).fetchone()["c"] >= s["group_size"]

        if all_ready:
            return "done"
        else:
            return "reveal"

    if r > 1:
        all_ready = con.execute(
            "SELECT COUNT(*) c FROM participants WHERE session_id=%s AND ready_for_next=1",
            (s["id"],)
        ).fetchone()["c"] >= s["group_size"]

        if not all_ready:
            return "reveal"

    decided = con.execute(
        "SELECT 1 FROM decisions WHERE participant_id=%s AND round_number=%s", (p["id"], r)
    ).fetchone()
    if not decided: return "round"

    ph = con.execute(
        "SELECT watch_ends_at FROM round_phases WHERE session_id=%s AND round_number=%s",
        (s["id"], r)
    ).fetchone()
    if not ph: return "wait"

    return "reveal"

def state_to_url(state: str) -> str:
    return {
        "lobby": url_for("lobby"),
        "round": url_for("round_view"),
        "wait": url_for("wait_view"),
        "reveal": url_for("reveal"),
        "feedback": url_for("feedback"),
        "done": url_for("done"),
    }[state]

def guard(expect_state: str):
    def deco(fn):
        @wraps(fn)
        def inner(*args, **kwargs):
            if not g.participant: return redirect(url_for("join"))
            con = db()
            p = con.execute("SELECT * FROM participants WHERE id=%s", (g.participant["id"],)).fetchone()
            s = con.execute("SELECT * FROM sessions WHERE id=%s", (p["session_id"],)).fetchone()
            st = current_state(con, p, s)
            if st != expect_state: return redirect(state_to_url(st))
            return fn(*args, **kwargs)
        return inner
    return deco


# -------------------- Round finalization (atomic) --------------------
def _finalize_round_atomic(con, sid: str, r: int, s: dict):
    cursor = con.cursor()

    try:
        con.begin()

        cursor.execute(
            "SELECT COUNT(*) as c FROM decisions WHERE session_id=%s AND round_number=%s",
            (sid, r)
        )
        decided = cursor.fetchone()["c"]

        if decided < s["group_size"]:
            con.rollback()
            return

        cursor.execute(
            "SELECT COUNT(*) as c FROM decisions WHERE session_id=%s AND round_number=%s AND total_cost IS NULL",
            (sid, r)
        )
        missing = cursor.fetchone()["c"]

        if missing <= 0:
            con.rollback()
            return

        cursor.execute(
            """SELECT d.id, d.participant_id, d.choice, p.ptype, p.join_number
               FROM decisions d JOIN participants p ON p.id=d.participant_id
               WHERE d.session_id=%s AND d.round_number=%s
               ORDER BY p.join_number""",
            (sid, r)
        )
        rows = cursor.fetchall()

        total_A = sum(1 for row in rows if row["choice"] == "A")
        N = s["group_size"]
        M = float(s["starting_balance"] or 500)

        for row in rows:
            did = row["id"]
            pid = row["participant_id"]
            choice = row["choice"]
            ptype = row["ptype"] or 1

            if choice == "A":
                cost = a_cost_for(ptype)
                others_A = max(0, total_A - 1)
                b_cost_round = None
            else:
                others_A = total_A
                cost = b_cost_adapt(ptype, others_A, N)
                b_cost_round = cost

            payout = max(M - float(cost), 0)

            cursor.execute(
                """UPDATE decisions
                   SET a_cost=%s, b_cost=%s, total_cost=%s,
                       payout=%s, base_payout=%s, others_A=%s, b_cost_round=%s, reveal=1
                   WHERE id=%s AND total_cost IS NULL""",
                (
                    cost if choice == "A" else None,
                    cost if choice == "B" else None,
                    cost,
                    payout,
                    M,
                    others_A,
                    b_cost_round,
                    did
                )
            )

            cursor.execute("UPDATE participants SET balance=%s WHERE id=%s", (payout, pid))

        cursor.execute(
            "UPDATE participants SET current_round = current_round + 1, ready_for_next = 0 WHERE session_id=%s AND current_round=%s",
            (sid, r)
        )

        now = utc_now()
        sec = int(s["watch_time"] or s["reveal_window"] or 5)
        cursor.execute(
            """REPLACE INTO round_phases
               (session_id,round_number,decision_ends_at,watch_ends_at,created_at)
               VALUES (%s,%s,%s,%s,%s)""",
            (sid, r, iso_utc(now), iso_utc(now + timedelta(seconds=sec)), iso_utc(now))
        )

        con.commit()

    except Exception:
        try:
            con.rollback()
        except Exception:
            pass
        raise
    finally:
        cursor.close()


# -------------------- Public --------------------
@app.route("/")
def index():
    if g.participant:
        con = db()
        p = con.execute("SELECT * FROM participants WHERE id=%s", (g.participant["id"],)).fetchone()
        s = con.execute("SELECT * FROM sessions WHERE id=%s", (p["session_id"],)).fetchone()
        return redirect(state_to_url(current_state(con, p, s)))
    return redirect(url_for("join"))

@app.route("/logout")
def logout():
    flask_session.pop("participant_id", None)
    return redirect(url_for("join"))

@app.route("/join", methods=["GET", "POST"])
def join():
    con = db()
    if request.method == "POST":
        code = request.form.get("code", "").strip().upper()
        p = con.execute("SELECT * FROM participants WHERE code=%s", (code,)).fetchone()
        if not p:
            return render_template("join.html", error="Code unbekannt.")
        if p["completed"]:
            return render_template("join.html", error="Dieser Code wurde bereits abgeschlossen. Bitte neuen Code verwenden.")
        now = iso_utc(utc_now())
        if not p["joined"]:
            nxt = con.execute(
                "SELECT COALESCE(MAX(join_number),0)+1 AS n FROM participants WHERE session_id=%s AND joined=1",
                (p["session_id"],)
            ).fetchone()["n"]
            ptype = p["ptype"] or ((nxt-1) % 6) + 1
            con.execute(
                "UPDATE participants SET joined=1, join_number=%s, ptype=%s, created_at=COALESCE(created_at, %s) WHERE id=%s",
                (nxt, ptype, now, p["id"])
            )
        else:
            if not p["ptype"]:
                cnt = con.execute(
                    "SELECT COUNT(*) c FROM participants WHERE session_id=%s AND ptype IS NOT NULL",
                    (p["session_id"],)
                ).fetchone()["c"]
                ptype = (cnt % 6) + 1
                con.execute("UPDATE participants SET ptype=%s WHERE id=%s", (ptype, p["id"]))
            con.execute("UPDATE participants SET joined=1 WHERE id=%s", (p["id"],))
        flask_session["participant_id"] = p["id"]
        flask_session.permanent = False
        con.commit()
        p2 = con.execute("SELECT * FROM participants WHERE id=%s", (p["id"],)).fetchone()
        s = con.execute("SELECT * FROM sessions WHERE id=%s", (p["session_id"],)).fetchone()
        return redirect(state_to_url(current_state(con, p2, s)))
    return render_template("join.html", error=None)

@app.route("/lobby")
@guard("lobby")
def lobby():
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (g.participant["session_id"],)).fetchone()
    joined = con.execute(
        "SELECT COUNT(*) c FROM participants WHERE session_id=%s AND joined=1",
        (s["id"],)
    ).fetchone()["c"]
    return render_template("lobby.html", session=s, participant=g.participant, joined=joined)

@app.get("/lobby_status")
def lobby_status():
    sid = request.args.get("session_id")
    pid = request.args.get("participant_id")
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not s:
        return jsonify({"err": "unknown_session"}), 404
    joined = con.execute(
        "SELECT COUNT(*) c FROM participants WHERE session_id=%s AND joined=1",
        (sid,)
    ).fetchone()["c"]

    reset = False
    if pid:
        p = con.execute("SELECT joined FROM participants WHERE id=%s", (pid,)).fetchone()
        if p and not p["joined"]:
            reset = True

    return jsonify({"joined": joined, "group_size": s["group_size"], "ready": joined >= s["group_size"], "reset": reset})

# ---------- Round ----------
@app.route("/round")
@guard("round")
def round_view():
    con = db()
    p = g.participant
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (p["session_id"],)).fetchone()
    r = p["current_round"]
    ptype = p["ptype"] or 1
    N = s["group_size"]

    a_cost_display = a_cost_for(ptype)
    others_max = max(1, N - 1)
    b_row_costs = [int(b_cost_adapt(ptype, k, N)) for k in range(1, others_max + 1)]
    b_list = [{"others": k, "cost": b_row_costs[k-1]} for k in range(1, others_max + 1)]

    return render_template(
        "round.html",
        session=s,
        round_number=r,
        N=N,
        a_cost_display=a_cost_display,
        b_list=b_list,
        others_max=others_max,
        base_payout=int(s["starting_balance"] or 500),
        balance_current=int(s["starting_balance"] or 500),
        participant=p
    )

@app.post("/choose")
def choose():
    if not g.participant:
        return ("No participant", 400)
    data = request.get_json() or {}
    choice = (data.get("choice") or "").upper()
    if choice not in ("A", "B"):
        return ("Invalid choice", 400)
    con = db()
    p = g.participant
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (p["session_id"],)).fetchone()
    r = p["current_round"]

    already = con.execute(
        "SELECT 1 FROM decisions WHERE participant_id=%s AND round_number=%s",
        (p["id"], r)
    ).fetchone()
    if already:
        return jsonify({"ok": True})

    con.execute(
        "INSERT INTO decisions (session_id, participant_id, round_number, choice, created_at) VALUES (%s,%s,%s,%s,%s)",
        (s["id"], p["id"], r, choice, iso_utc(utc_now())),
    )
    con.commit()
    return jsonify({"ok": True})

@app.route("/wait")
@guard("wait")
def wait_view():
    con = db()
    p = g.participant
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (p["session_id"],)).fetchone()
    r = p["current_round"]
    decided = con.execute(
        "SELECT COUNT(*) c FROM decisions WHERE session_id=%s AND round_number=%s",
        (s["id"], r)
    ).fetchone()["c"]
    return render_template("wait.html", session=s, round_number=r, decided=decided, participant=p)

@app.get("/round_status")
def round_status():
    sid = request.args.get("session_id")
    r = int(request.args.get("round"))
    pid = request.args.get("participant_id")
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not s:
        return jsonify({"err": "unknown_session"}), 404

    reset = False
    if pid:
        p = con.execute("SELECT joined FROM participants WHERE id=%s", (pid,)).fetchone()
        if p and not p["joined"]:
            reset = True
    if reset:
        return jsonify({"reset": True})

    decided = con.execute(
        "SELECT COUNT(*) c FROM decisions WHERE session_id=%s AND round_number=%s",
        (sid, r)
    ).fetchone()["c"]
    ready = decided >= s["group_size"]

    players_payload = []
    watch_ends_at = None

    if ready:
        try:
            _finalize_round_atomic(con, sid, r, s)
        except pymysql.OperationalError:
            pass

        rp = con.execute(
            "SELECT * FROM round_phases WHERE session_id=%s AND round_number=%s",
            (sid, r)
        ).fetchone()
        watch_ends_at = rp["watch_ends_at"] if rp else None

        for row in con.execute("""
             SELECT p.join_number, d.choice, d.total_cost, d.payout
             FROM decisions d JOIN participants p ON p.id=d.participant_id
             WHERE d.session_id=%s AND d.round_number=%s ORDER BY p.join_number
        """, (sid, r)).fetchall():
            players_payload.append({
                "player_no": row["join_number"],
                "choice": row["choice"],
                "cost": row["total_cost"],
                "payout": row["payout"],
            })

    decided_players = [row["join_number"] for row in con.execute(
        "SELECT p.join_number FROM decisions d JOIN participants p ON p.id=d.participant_id "
        "WHERE d.session_id=%s AND d.round_number=%s ORDER BY p.join_number",
        (sid, r)
    ).fetchall()]

    return jsonify({
        "decided": decided,
        "ready": ready,
        "decided_players": decided_players,
        "watch_ends_at": watch_ends_at,
        "players": players_payload
    })

# ---------- Reveal ----------
@app.route("/reveal")
@guard("reveal")
def reveal():
    con = db()
    p = g.participant
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (p["session_id"],)).fetchone()
    r = p["current_round"] - 1
    if r < 1: return redirect(url_for("round_view"))
    is_last_round = (p["current_round"] > s["rounds"])
    return render_template("reveal.html", session=s, round_number=r, participant=p, is_last_round=is_last_round)

@app.get("/reveal_status")
def reveal_status():
    sid = request.args.get("session_id")
    r = int(request.args.get("round") or 0)
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not s or r < 1: return jsonify({"err":"bad"}), 400

    ph = con.execute(
        "SELECT decision_ends_at, watch_ends_at FROM round_phases WHERE session_id=%s AND round_number=%s",
        (sid, r)
    ).fetchone()
    now = utc_now()
    if not ph:
        sec = int(s["reveal_window"] or 5)
        con.execute(
            "REPLACE INTO round_phases (session_id,round_number,decision_ends_at,watch_ends_at,created_at) VALUES (%s,%s,%s,%s,%s)",
            (sid, r, iso_utc(now), iso_utc(now + timedelta(seconds=sec)), iso_utc(now))
        )
        con.commit()
        ends_at = iso_utc(now + timedelta(seconds=sec))
    else:
        ends_at = ph["watch_ends_at"] if ph["watch_ends_at"].endswith("Z") else ph["watch_ends_at"] + "Z"

    con.execute(
        "UPDATE decisions SET reveal=1 WHERE session_id=%s AND round_number=%s AND (reveal IS NULL OR reveal!=1)",
        (sid, r)
    )
    con.commit()

    rows = con.execute("""
        SELECT p.id as pid, p.code, p.join_number, d.choice, d.payout
        FROM participants p
        LEFT JOIN decisions d ON d.participant_id=p.id AND d.round_number=%s
        WHERE p.session_id=%s
        ORDER BY p.join_number, p.code
    """, (r, sid)).fetchall()

    players = []
    me = None
    for row in rows:
        obj = {
            "code": row["code"],
            "player_no": row["join_number"],
            "choice": row["choice"],
            "payout": row["payout"],
        }
        players.append(obj)
        if g.participant and row["pid"] == g.participant["id"]:
            me = obj

    ph2 = con.execute(
        "SELECT watch_ends_at FROM round_phases WHERE session_id=%s AND round_number=%s",
        (sid, r)
    ).fetchone()
    phase = "watch"
    if ph2 and utc_now() >= parse_iso_utc(ph2["watch_ends_at"]):
        phase = "done"
        ends_at = iso_utc(utc_now())

    return jsonify({"phase": phase, "ends_at": ends_at, "total": len(players), "players": players, "me": me})

# ---------- Ready Confirmation ----------
@app.post("/confirm_ready")
def confirm_ready():
    """Player confirms they are ready for the next round."""
    if not g.participant:
        return ("No participant", 400)
    con = db()
    p = g.participant
    con.execute("UPDATE participants SET ready_for_next=1 WHERE id=%s", (p["id"],))
    con.commit()
    return jsonify({"ok": True})

@app.get("/ready_status")
def ready_status():
    """Returns status of who is ready for the next round."""
    sid = request.args.get("session_id")
    pid = request.args.get("participant_id")
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not s:
        return jsonify({"err": "unknown_session"}), 404

    reset = False
    if pid:
        p = con.execute("SELECT joined FROM participants WHERE id=%s", (pid,)).fetchone()
        if p and not p["joined"]:
            reset = True
    if reset:
        return jsonify({"reset": True})

    rows = con.execute(
        """SELECT p.id, p.join_number, p.ready_for_next
           FROM participants p WHERE p.session_id=%s ORDER BY p.join_number""",
        (sid,)
    ).fetchall()

    ready_count = sum(1 for r in rows if r["ready_for_next"])
    all_ready = ready_count >= s["group_size"]

    players = []
    me_ready = False
    for row in rows:
        is_ready = bool(row["ready_for_next"])
        players.append({
            "player_no": row["join_number"],
            "ready": is_ready
        })
        if g.participant and row["id"] == g.participant["id"]:
            me_ready = is_ready

    return jsonify({
        "ready_count": ready_count,
        "group_size": s["group_size"],
        "all_ready": all_ready,
        "me_ready": me_ready,
        "players": players
    })

# ---------- Feedback ----------
@app.route("/feedback")
@guard("feedback")
def feedback():
    con = db()
    p = g.participant
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (p["session_id"],)).fetchone()
    r = p["current_round"] - 1
    if r < 1:
        return redirect(url_for("round_view"))

    d = con.execute(
        "SELECT choice, total_cost, payout, base_payout, b_cost_round, others_A "
        "FROM decisions WHERE session_id=%s AND participant_id=%s AND round_number=%s",
        (s["id"], p["id"], r),
    ).fetchone()

    decided_A = con.execute(
        "SELECT COUNT(*) AS c FROM decisions WHERE session_id=%s AND round_number=%s AND choice='A'",
        (s["id"], r),
    ).fetchone()["c"]
    decided_B = con.execute(
        "SELECT COUNT(*) AS c FROM decisions WHERE session_id=%s AND round_number=%s AND choice='B'",
        (s["id"], r),
    ).fetchone()["c"]

    ctx = dict(
        session=s,
        N=s["group_size"],
        round_number=r,
        my_choice=d["choice"] if d else None,
        my_cost=d["total_cost"] if d else None,
        payout=(d["payout"] if d else None),
        base_payout=(d["base_payout"] if d else s["starting_balance"]),
        b_cost_round=(d["b_cost_round"] if d else None),
        others_A=(d["others_A"] if d else None),
        decided_A=decided_A,
        decided_B=decided_B,
        next_round=(not s["archived"]) and (p["current_round"] <= s["rounds"]),
    )
    return render_template("feedback.html", **ctx)

@app.route("/done")
@guard("done")
def done():
    con = db()
    pid = flask_session.get("participant_id")
    balance = None
    code = None
    if pid:
        row = con.execute("SELECT code, balance FROM participants WHERE id=%s", (pid,)).fetchone()
        if row:
            balance = row["balance"]
            code = row["code"]
            con.execute("UPDATE participants SET completed=1 WHERE id=%s", (pid,))
            con.commit()
    flask_session.pop("participant_id", None)
    return render_template("done.html", balance=balance, code=code)

@app.get("/healthz")
def healthz():
    return "ok", 200


# -------------------- Admin --------------------
def require_admin():
    return bool(flask_session.get("admin_ok"))

def _session_done(con, sid):
    row = con.execute("SELECT group_size, rounds FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not row:
        return False
    grp = row["group_size"]
    rmax = row["rounds"]
    cnt = con.execute(
        "SELECT COUNT(*) c FROM participants WHERE session_id=%s AND current_round > %s",
        (sid, rmax)
    ).fetchone()["c"]
    return cnt >= grp

@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            flask_session["admin_ok"] = True
            return redirect(url_for("admin"))
        return render_template("admin_login.html", error="Falsches Passwort.", admin_tab_guard=True)
    return render_template("admin_login.html", error=None, admin_tab_guard=True)

@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not require_admin():
        return redirect(url_for("admin_login"))
    con = db()

    if request.method == "POST":
        name = request.form.get("name", f"Session {datetime.datetime.now():%Y-%m-%d %H:%M}")
        group_size = int(request.form.get("group_size", "6"))
        rounds = int(request.form.get("rounds", "20"))
        base_payout = int(request.form.get("base_payout", "500"))

        cvac = 0.0
        alpha = 0.0
        cinf = 0.0
        subsidy = 0
        subsidy_amount = 0.0
        cost_mode = "type_table"

        sid = str(uuid.uuid4())
        con.execute("""
            INSERT INTO sessions
              (id,name,group_size,rounds,cvac,alpha,cinf,subsidy,subsidy_amount,
               starting_balance,created_at,archived,reveal_window,watch_time,cost_mode)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            sid, name, group_size, rounds, cvac, alpha, cinf, subsidy, subsidy_amount,
            base_payout, iso_utc(utc_now()), 0, 5, 5, cost_mode
        ))

        for i in range(group_size):
            pid = str(uuid.uuid4())
            # Generate unique code (max 20 attempts to prevent infinite loops)
            code = None
            for attempt in range(20):
                code = create_code(6)
                if not con.execute("SELECT 1 FROM participants WHERE code=%s", (code,)).fetchone():
                    break
            else:
                # Should never happen with 6-char codes (32^6 = ~1B possibilities)
                raise Exception(f"Failed to generate unique participant code after 20 attempts")
            ptype = (i % 6) + 1
            theta = 0.0
            lambd = 0.0
            con.execute(
                "INSERT INTO participants (id,session_id,code,theta,lambda,joined,join_number,current_round,balance,completed,created_at,ptype) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (pid, sid, code, theta, lambd, 0, None, 1, base_payout, 0, iso_utc(utc_now()), ptype)
            )
        con.commit()
        return redirect(url_for("admin"))

    rows = con.execute("SELECT * FROM sessions ORDER BY created_at DESC").fetchall()
    sessions_active, sessions_done, sessions_arch = [], [], []
    for s in rows:
        ps = con.execute("SELECT code FROM participants WHERE session_id=%s", (s["id"],)).fetchall()
        sdict = {**dict(s), "participants": [dict(p) for p in ps]}
        if s["archived"]:
            sessions_arch.append(sdict)
        else:
            if _session_done(con, s["id"]):
                sessions_done.append(sdict)
            else:
                sessions_active.append(sdict)

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    return render_template(
        "admin.html",
        sessions_active=sessions_active,
        sessions_done=sessions_done,
        sessions_arch=sessions_arch,
        now=now,
        admin_tab_guard=True
    )

@app.get("/admin/session/<session_id>")
def admin_session_view(session_id):
    if not require_admin():
        return redirect(url_for("admin_login"))
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (session_id,)).fetchone()
    if not s:
        return redirect(url_for("admin"))
    r = con.execute(
        "SELECT MIN(current_round) AS r FROM participants WHERE session_id=%s",
        (session_id,)
    ).fetchone()["r"] or 1
    r = min(r, s["rounds"])
    return render_template("admin_session.html", session=s, round_number=r, admin_tab_guard=True)

@app.get("/admin/session_status")
def admin_session_status():
    if not require_admin():
        return ("Forbidden", 403)
    sid = request.args.get("session_id")
    con = db()
    srow = con.execute("SELECT * FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not srow:
        return jsonify({"participants": [], "decided_count": 0, "session": None})

    r = con.execute(
        "SELECT MIN(current_round) AS r FROM participants WHERE session_id=%s",
        (sid,)
    ).fetchone()["r"] or 1
    r_disp = min(r, srow["rounds"])

    rows = con.execute(
        """SELECT p.id, p.code, p.join_number, p.balance, p.current_round, p.ready_for_next,
                  EXISTS(SELECT 1 FROM decisions d WHERE d.participant_id=p.id AND d.round_number=%s) AS decided,
                  (SELECT d.choice FROM decisions d WHERE d.participant_id=p.id AND d.round_number=%s LIMIT 1) AS choice
           FROM participants p WHERE p.session_id=%s ORDER BY p.join_number, p.code""",
        (r, r, sid)
    ).fetchall()

    participants = [{
        "id": rr["id"],
        "code": rr["code"],
        "player_no": rr["join_number"],
        "balance": rr["balance"],
        "round_display": min(rr["current_round"], srow["rounds"]),
        "decided": bool(rr["decided"]),
        "choice": rr["choice"],
        "ready_for_next": bool(rr["ready_for_next"])
    } for rr in rows]

    decided_count = sum(1 for x in participants if x["decided"])
    ready_count = sum(1 for x in participants if x["ready_for_next"])
    return jsonify({
        "participants": participants,
        "decided_count": decided_count,
        "ready_count": ready_count,
        "session": {"id": srow["id"], "current_round": r_disp}
    })

@app.post("/admin/reset_session")
def admin_reset_session():
    if not require_admin():
        return redirect(url_for("admin_login"))
    sid = request.form.get("session_id")
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not s:
        return redirect(url_for("admin"))

    con.execute("DELETE FROM decisions WHERE session_id=%s", (sid,))
    con.execute("DELETE FROM round_phases WHERE session_id=%s", (sid,))
    con.execute(
        "UPDATE participants SET current_round=1, join_number=NULL, joined=0, balance=%s, completed=0, ready_for_next=0 WHERE session_id=%s",
        (s["starting_balance"], sid)
    )
    con.commit()
    con.execute("UPDATE sessions SET archived=0 WHERE id=%s", (sid,))
    con.commit()
    return redirect(url_for("admin"))

@app.post("/admin/archive_session")
def admin_archive_session():
    if not require_admin():
        return redirect(url_for("admin_login"))
    sid = request.form.get("session_id")
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not s:
        return redirect(url_for("admin"))

    ensure_archive_schema(con, "sessions")
    ensure_archive_schema(con, "participants")
    ensure_archive_schema(con, "decisions")

    con.execute("START TRANSACTION")
    con.execute("INSERT INTO archived_sessions SELECT * FROM sessions WHERE id=%s", (sid,))
    con.execute("INSERT INTO archived_participants SELECT * FROM participants WHERE session_id=%s", (sid,))
    con.execute("INSERT INTO archived_decisions SELECT * FROM decisions WHERE session_id=%s", (sid,))
    con.execute("UPDATE sessions SET archived=1 WHERE id=%s", (sid,))
    con.execute("UPDATE participants SET completed=1 WHERE session_id=%s", (sid,))
    con.commit()
    return redirect(url_for("admin"))

@app.post("/admin/delete_session")
def admin_delete_session():
    if not require_admin():
        return redirect(url_for("admin_login"))
    sid = request.form.get("session_id")
    con = db()
    exists = con.execute("SELECT 1 FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not exists:
        return redirect(url_for("admin"))

    con.execute("START TRANSACTION")
    con.execute("DELETE FROM decisions WHERE session_id=%s", (sid,))
    con.execute("DELETE FROM round_phases WHERE session_id=%s", (sid,))
    con.execute("DELETE FROM participants WHERE session_id=%s", (sid,))
    con.execute("DELETE FROM sessions WHERE id=%s", (sid,))
    con.commit()
    return redirect(url_for("admin"))

# --------- XLSX Export ----------
def _style_table(ws, header_row=1, wrap_cols=None, int_cols=None):
    hdr_fill = PatternFill("solid", fgColor="1F2A44")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = f"A{header_row+1}"
    ws.auto_filter.ref = ws.dimensions

    if int_cols:
        for col_idx in int_cols:
            col_letter = get_column_letter(col_idx)
            for rr in range(header_row+1, ws.max_row+1):
                ws[f"{col_letter}{rr}"].number_format = "0"

    if wrap_cols:
        for col_idx in wrap_cols:
            col_letter = get_column_letter(col_idx)
            for rr in range(header_row, ws.max_row+1):
                ws[f"{col_letter}{rr}"].alignment = Alignment(wrap_text=True, vertical="top")

    max_width = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            length = len(str(cell.value)) if cell.value is not None else 0
            max_width[cell.column] = max(max_width.get(cell.column, 0), length)

    for col, width in max_width.items():
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = min(60, max(10, width * 1.15))

@app.get("/admin/export_session_xlsx")
def admin_export_session_xlsx():
    if not require_admin():
        return redirect(url_for("admin_login"))
    sid = request.args.get("session_id")
    con = db()
    s = con.execute("SELECT * FROM sessions WHERE id=%s", (sid,)).fetchone()
    if not s:
        return ("Not found", 404)

    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Session"
    ws0.append(["id","name","group_size","rounds","starting_balance","created_at","archived"])
    ws0.append([
        s["id"], s["name"], s["group_size"], s["rounds"],
        s["starting_balance"], s["created_at"], s["archived"]
    ])
    _style_table(ws0, header_row=1, wrap_cols=[6,7], int_cols=[3,4,5])

    ws1 = wb.create_sheet("Participants")
    ws1.append(["player_no","code","ptype","joined","current_round","balance","completed","ready_for_next","created_at"])
    for p in con.execute(
        "SELECT join_number, code, ptype, joined, current_round, balance, completed, ready_for_next, created_at "
        "FROM participants WHERE session_id=%s ORDER BY join_number, code",
        (sid,)
    ):
        ws1.append([p["join_number"], p["code"], p["ptype"], p["joined"],
                    p["current_round"], p["balance"], p["completed"], p["ready_for_next"], p["created_at"]])
    _style_table(ws1, header_row=1, wrap_cols=[9], int_cols=[1,3,4,5,6,7,8])

    ws2 = wb.create_sheet("Decisions")
    ws2.append(["round","player_no","code","ptype","choice","a_cost","b_cost","total_cost",
                "payout","created_at","revealed","others_A","b_cost_round","base_payout"])
    for d in con.execute("""
        SELECT d.round_number, p.join_number, p.code, p.ptype, d.choice,
               d.a_cost, d.b_cost, d.total_cost, d.payout, d.created_at, d.reveal,
               d.others_A, d.b_cost_round, d.base_payout
        FROM decisions d JOIN participants p ON p.id=d.participant_id
        WHERE d.session_id=%s ORDER BY d.round_number, p.join_number, p.code
    """, (sid,)):
        ws2.append([d["round_number"], d["join_number"], d["code"], d["ptype"], d["choice"],
                    d["a_cost"], d["b_cost"], d["total_cost"], d["payout"], d["created_at"], d["reveal"],
                    d["others_A"], d["b_cost_round"], d["base_payout"]])
    _style_table(ws2, header_row=1, wrap_cols=[10], int_cols=[1,2,4,6,7,8,9,11,12,13,14])

    ws3 = wb.create_sheet("Design")
    ws3.append(["Parameter","Wert","Kommentar"])
    for k, v, c in [
        ("Session ID", s["id"], ""),
        ("Session Name", s["name"], ""),
        ("Gruppengroesse (N)", s["group_size"], "Anzahl Teilnehmende pro Gruppe"),
        ("Runden", s["rounds"], "Anzahl Perioden; Parameter konstant"),
        ("Basisbetrag M", s["starting_balance"], "Rundenstart; Auszahlung = M - Kosten"),
        ("Erstellt (UTC)", s["created_at"], ""),
        ("Archiviert", s["archived"], "1 = archiviert"),
    ]:
        ws3.append([k, v, c])
    _style_table(ws3, header_row=1, wrap_cols=[2,3])

    ws4 = wb.create_sheet("TypeCostTable")
    ws4.append(["Typ","A_cost","B_cost_1A","B_cost_2A","B_cost_3A","B_cost_4A","B_cost_5A"])
    for t in sorted(TYPE_COST.keys()):
        ws4.append([t, TYPE_COST[t]["A"], *TYPE_COST[t]["B"][:5]])
    _style_table(ws4, header_row=1, int_cols=[1,2,3,4,5,6,7])

    ws5 = wb.create_sheet("RoundSettings")
    ws5.append(["round","M","N"])
    for rr in range(1, int(s["rounds"]) + 1):
        ws5.append([rr, s["starting_balance"], s["group_size"]])
    _style_table(ws5, header_row=1, int_cols=[1,2,3])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"session_{s['name'].replace(' ', '_')}_{s['id'][:8]}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# -------------------- Run --------------------
if __name__ == "__main__":
    init_db()
    app.run(host="127.0.0.1", port=5000, debug=DEBUG_MODE)
